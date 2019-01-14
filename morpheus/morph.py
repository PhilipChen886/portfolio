import sys
import io
from openpyxl import load_workbook

sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

origin = ""

def import_text(path):
    wb = load_workbook(path)
    sheet = wb['Sheet1']
    global origin
    origin = sheet.cell(row=1, column=1).value
    return

def print_origin():
    print(origin)
    return

def parse():                                                #!NOTE  hierachy: paragraph >> sentence >> sub-sentence >> pre-elements >> elements
    global main_list
    main_list = origin.splitlines()                         # removes newlines
    for i in range (len (main_list)-1, -1, -1):             # removes blank elements
        if (main_list[i] == ''):
            main_list.pop(i)
            i = i+1
    return

def new_parse():
    global main_list2
    main_list2 = [ [""], [""] ]
    j = 0
    k = 0
    #char_counter = 0
    for i in range( len(origin) ):
        #!NOTE character check

        #!NOTE parsing
        if (origin[i] == "\n" and origin[i+1] == "\n"):
            main_list2.append( [""] )
            j += 1
            k = 0
        elif (origin[i-1] == "。" and i != 0):
            main_list2[j].append( "" )
            k += 1

        if (origin[i] != "\n"):
            main_list2[j][k] += origin[i]
        #char_counter += 1
    #print(char_counter)
    #print(sys.getsizeof(main_list2))
    return

def print_main():
    for i in range (len (main_list) ):
        print(main_list[i])
    return

def print_main_list():
    print(main_list)
    return

def print_test():
    for i in range(12353, 12436):
        print( chr(i) )